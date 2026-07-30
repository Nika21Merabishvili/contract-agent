"""Turn the analyzer's JSON into an Excel workbook -- step 2 of 2.

Step 1 (`pdf_analyze.py`) is "PDF -> JSON"; this is "JSON -> xlsx". The two stay
separate programs so a run can be inspected, kept, or re-exported without paying
for the model again:

    python pdf_analyze.py contract.pdf | python export_excel.py -o out.xlsx
    python pdf_analyze.py contract.pdf > result.json
    python export_excel.py result.json                # same sheet, no model call

One workbook, one worksheet, one table: row 1 is the Georgian header, and every
analysis below it is one row. Columns come only from `contract_data` and
`tax_analysis`. `_audit` is a reasoning trace for scoring the model, not a thing
end users read, so it never reaches the sheet. The one exception is a trailing
note column that appears only when a contract was OCR'd from a scanned PDF --
see OCR_COLUMN_LABEL below.

The headers are not written here -- they come from `georgian.FIELD_LABELS`, next
to the rest of the app's Georgian, so the sheet cannot drift out of sync with it.
Values are copied verbatim: they were already turned into Georgian upstream.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import diagnostics as diag
import georgian as ka
from georgian import NOT_SPECIFIED
from schemas import CONTRACT_FIELDS, FREE_TEXT_FIELDS, TAX_FIELDS

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit(
        "error: openpyxl is not installed.\n  pip install openpyxl"
    ) from None


# The two blocks that reach the sheet, in column order. `_audit` is absent by
# design -- see the module docstring.
BLOCKS = (("contract_data", CONTRACT_FIELDS), ("tax_analysis", TAX_FIELDS))

SHEET_TITLE = "ხელშეკრულებები"

FONT_NAME = "Arial"
HEADER_FILL = "1F4E78"
HEADER_TEXT = "FFFFFF"

# Free text runs to several sentences; everything else is a name, a code, a date
# or a number. The wide/wrapped columns are exactly the fields the model wrote
# prose for, so the list is imported rather than re-guessed here.
WIDE_WIDTH = 55
DEFAULT_WIDTH = 24

# The optional leading column a batch export adds -- see build_workbook(sources=...).
SOURCE_COLUMN_LABEL = "წყარო ფაილი"

# The optional trailing column marking OCR-derived rows. Only added when at
# least one analysis carries `"_source": "ocr"` (a scanned contract that went
# through the OCR fallback), so the validated sheet layout is untouched for
# every normal run. The note names the verbatim fields worth re-checking: OCR
# can misread exactly the characters -- names, ID numbers, amounts, dates --
# that this sheet exists to copy faithfully.
OCR_COLUMN_LABEL = "შენიშვნა"
OCR_NOTE = (
    "დასკანერებული დოკუმენტი (OCR) — გადაამოწმეთ სახელები, "
    "საიდენტიფიკაციო ნომრები, თანხები და თარიღები"
)


# --------------------------------------------------------------------------- #
# Input
# --------------------------------------------------------------------------- #


def load_analyses(path: Path | None) -> list[dict]:
    """Read the analyzer's JSON from a file or from piped stdin.

    Accepts a single analysis object or a list of them, so several contracts
    become several rows in one sheet.
    """
    if path is not None:
        if not path.is_file():
            raise SystemExit(f"error: no such file: {path}")
        raw = path.read_text(encoding="utf-8")
        source = str(path)
    else:
        if sys.stdin is None or sys.stdin.isatty():
            raise SystemExit(
                "error: no input.\n"
                "  python export_excel.py result.json\n"
                "  python pdf_analyze.py contract.pdf | python export_excel.py"
            )
        raw = sys.stdin.read()
        source = "stdin"

    if not raw.strip():
        raise SystemExit(f"error: {source} is empty -- nothing to export.")

    try:
        data = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise SystemExit(f"error: {source} is not valid JSON: {exc}") from None

    analyses = data if isinstance(data, list) else [data]
    for index, item in enumerate(analyses, start=1):
        if not isinstance(item, dict):
            raise SystemExit(
                f"error: {source}: entry {index} is {type(item).__name__}, "
                "expected an analysis object."
            )
    if not analyses:
        raise SystemExit(f"error: {source} holds an empty list -- nothing to export.")
    return analyses


# --------------------------------------------------------------------------- #
# Columns
# --------------------------------------------------------------------------- #


def columns(analyses: list[dict]) -> list[tuple[str, str]]:
    """The (block, key) pair behind each column, left to right.

    The canonical field lists come first and always in full, so two runs of this
    script produce the same columns in the same order even if one contract's JSON
    is short a key. Anything present in the data but not in those lists is
    appended rather than dropped -- a field added upstream should show up in the
    sheet (under its raw key, which is visibly wrong and easy to fix) instead of
    silently disappearing.
    """
    pairs = [(block, key) for block, fields in BLOCKS for key in fields]
    known = set(pairs)
    for analysis in analyses:
        for block, _ in BLOCKS:
            for key in analysis.get(block) or {}:
                if (block, key) not in known:
                    known.add((block, key))
                    pairs.append((block, key))
    return pairs


def cell_value(analysis: dict, block: str, key: str):
    """One cell. A missing key is data, not an error.

    The pipeline already writes the sentinel for a fact the contract is silent
    on; a key missing outright means the same thing to a reader, so it is shown
    the same way and the row stays complete and aligned.
    """
    value = (analysis.get(block) or {}).get(key)
    if value is None or (isinstance(value, str) and not value.strip()):
        return NOT_SPECIFIED
    if isinstance(value, bool):  # before int -- bool is an int in Python
        return str(value)
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


# --------------------------------------------------------------------------- #
# Output
# --------------------------------------------------------------------------- #


def build_workbook(analyses: list[dict], *, sources: list[str] | None = None) -> Workbook:
    """Build the one-sheet workbook in memory and return it, unsaved.

    Split out from `write_workbook` so a caller that wants the bytes rather than a
    file on disk can save the returned book anywhere openpyxl accepts, including a
    `BytesIO` -- the web app streams the .xlsx straight back to the browser and
    never touches the filesystem for it. `write_workbook` is the file-writing
    wrapper around this, and the CLI still goes through it unchanged.

    `sources`, if given, is one label per analysis (same order -- typically the
    contract's filename) and adds a leading column identifying each row, so a
    multi-contract batch stays legible on one sheet. It is None by default: a
    single-contract call -- the CLI, or one contract analysed on its own -- gets
    exactly the sheet it always has, with no extra column.
    """
    if sources is not None and len(sources) != len(analyses):
        raise ValueError(f"sources has {len(sources)} entries for {len(analyses)} analyses")

    pairs = columns(analyses)
    offset = 1 if sources is not None else 0

    # See OCR_COLUMN_LABEL: a trailing note column, present only when some row
    # actually came through OCR.
    ocr_flags = [analysis.get("_source") == "ocr" for analysis in analyses]
    ocr_column = (len(pairs) + offset + 1) if any(ocr_flags) else None

    book = Workbook()
    sheet = book.active
    sheet.title = SHEET_TITLE

    header_font = Font(name=FONT_NAME, size=11, bold=True, color=HEADER_TEXT)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(name=FONT_NAME, size=11)
    edge = Side(style="thin", color="BFBFBF")
    body_border = Border(left=edge, right=edge, top=edge, bottom=edge)

    if sources is not None:
        cell = sheet.cell(row=1, column=1, value=SOURCE_COLUMN_LABEL)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for column, (_, key) in enumerate(pairs, start=1 + offset):
        cell = sheet.cell(row=1, column=column, value=ka.field_label(key))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    if ocr_column is not None:
        cell = sheet.cell(row=1, column=ocr_column, value=OCR_COLUMN_LABEL)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    sheet.row_dimensions[1].height = 46

    for row, analysis in enumerate(analyses, start=2):
        if sources is not None:
            cell = sheet.cell(row=row, column=1, value=sources[row - 2])
            cell.font = body_font
            cell.border = body_border
            cell.alignment = Alignment(vertical="top")
        for column, (block, key) in enumerate(pairs, start=1 + offset):
            cell = sheet.cell(row=row, column=column, value=cell_value(analysis, block, key))
            cell.font = body_font
            cell.border = body_border
            cell.alignment = Alignment(
                vertical="top", wrap_text=key in FREE_TEXT_FIELDS
            )
        if ocr_column is not None:
            cell = sheet.cell(
                row=row, column=ocr_column, value=OCR_NOTE if ocr_flags[row - 2] else ""
            )
            cell.font = body_font
            cell.border = body_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if sources is not None:
        sheet.column_dimensions["A"].width = DEFAULT_WIDTH
    for column, (_, key) in enumerate(pairs, start=1 + offset):
        letter = get_column_letter(column)
        sheet.column_dimensions[letter].width = (
            WIDE_WIDTH if key in FREE_TEXT_FIELDS else DEFAULT_WIDTH
        )
    if ocr_column is not None:
        sheet.column_dimensions[get_column_letter(ocr_column)].width = WIDE_WIDTH

    last = get_column_letter(ocr_column or (len(pairs) + offset))
    sheet.freeze_panes = "B2" if sources is not None else "A2"
    sheet.auto_filter.ref = f"A1:{last}{len(analyses) + 1}"

    return book


def write_workbook(
    analyses: list[dict], out_path: Path, *, sources: list[str] | None = None
) -> Path:
    """Write one sheet to `out_path`: a Georgian header row and one row per analysis."""
    book = build_workbook(analyses, sources=sources)

    out_path = out_path.expanduser()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        book.save(out_path)
    except PermissionError:
        raise SystemExit(
            f"error: cannot write {out_path} -- it is open in Excel. Close it and re-run."
        ) from None
    return out_path


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #


def default_out(path: Path | None) -> Path:
    """Beside the JSON when there is one; otherwise in the working directory."""
    return path.with_suffix(".xlsx") if path is not None else Path("contracts.xlsx")


def main() -> None:
    # Same reason as cli.py: the Windows console (and a piped stdin) default to
    # cp1252, which mangles or raises on Georgian.
    for stream in (sys.stdin, sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8", errors="replace")

    parser = argparse.ArgumentParser(
        description="Turn pdf_analyze.py's JSON into an Excel workbook (step 2 of 2).",
        epilog="reads piped stdin when no JSON file is given",
    )
    parser.add_argument(
        "json_file", type=Path, nargs="?", help="analyzer JSON; reads stdin if omitted"
    )
    parser.add_argument(
        "-o", "--out", type=Path,
        help="where to write the .xlsx (default: beside the JSON, else ./contracts.xlsx)",
    )
    args = parser.parse_args()

    analyses = load_analyses(args.json_file)
    out_path = write_workbook(analyses, args.out or default_out(args.json_file))

    rows = len(analyses)
    diag.progress(f"Saved: {out_path}  ({rows} contract{'' if rows == 1 else 's'})")


if __name__ == "__main__":
    main()
